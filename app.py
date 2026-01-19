import streamlit as st
import requests
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
import io
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import base64

# ========== PAGE CONFIG ==========
st.set_page_config(
    page_title="UXXCA - AI CFO Assistant",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ========== FIXED THEME - GUARANTEED WHITE TEXT ==========
st.markdown("""
<style>
    /* FORCE DARK BACKGROUND EVERYWHERE */
    .stApp {
        background-color: #0f172a !important;
    }
    
    /* FORCE ALL TEXT TO BE WHITE */
    * {
        color: #ffffff !important;
    }
    
    /* Specific fixes for Streamlit components */
    .stMarkdown, .stMarkdown p, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4, .stMarkdown h5, .stMarkdown h6 {
        color: #ffffff !important;
    }
    
    /* Fix metrics */
    [data-testid="stMetricValue"], [data-testid="stMetricLabel"] {
        color: #ffffff !important;
    }
    
    /* Fix dataframes */
    .stDataFrame {
        color: #ffffff !important;
    }
    
    /* Fix input labels */
    .stTextInput label, .stNumberInput label, .stTextArea label {
        color: #ffffff !important;
    }
    
    /* UXXCA Brand Header */
    .uxxca-header {
        background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%);
        padding: 2.5rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        color: white !important;
        box-shadow: 0 10px 30px rgba(79, 70, 229, 0.3);
    }
    
    /* Metric Cards */
    .metric-card {
        background: #1e293b;
        padding: 1.5rem;
        border-radius: 12px;
        border-left: 5px solid #60a5fa;
        margin: 0.5rem 0;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    
    /* Chat Input Fix */
    .stChatInput {
        background-color: #1e293b !important;
        border-radius: 15px !important;
        border: 2px solid #4f46e5 !important;
        margin-top: 20px !important;
    }
    
    .stChatInput input {
        background-color: transparent !important;
        color: #ffffff !important;
        font-size: 1rem !important;
        padding: 0.8rem !important;
    }
    
    .stChatInput input::placeholder {
        color: #94a3b8 !important;
    }
    
    /* Graph Container */
    .graph-box {
        background-color: #1e293b;
        border-radius: 15px;
        padding: 1.5rem;
        margin: 1rem 0;
        border: 1px solid #334155;
    }
</style>
""", unsafe_allow_html=True)

# ========== PROFESSIONAL SPREADSHEET GENERATOR ==========
def create_financial_spreadsheet(financial_data, transactions=None):
    """Create a professional Excel spreadsheet with financial data"""
    
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Financial Dashboard"
    
    # Title
    ws1.merge_cells('A1:F1')
    title_cell = ws1['A1']
    title_cell.value = "UXXCA FINANCIAL DASHBOARD"
    title_cell.font = Font(size=20, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center')
    
    # Company Info
    ws1['A3'] = "Company Name:"
    ws1['B3'] = financial_data.get('company_name', 'Your Business')
    ws1['A4'] = "Report Period:"
    ws1['B4'] = datetime.now().strftime("%B %Y")
    
    # Key Metrics
    metrics = [
        ["Monthly Revenue", f"${financial_data['revenue']:,.2f}"],
        ["Monthly Expenses", f"${financial_data['expenses']:,.2f}"],
        ["Monthly Profit", f"${financial_data['revenue'] - financial_data['expenses']:,.2f}"],
        ["Profit Margin", f"{(financial_data['revenue'] - financial_data['expenses']) / financial_data['revenue'] * 100:.1f}%"],
        ["Cash Balance", f"${financial_data['cash_balance']:,.2f}"],
        ["Runway", f"{financial_data['cash_balance'] / financial_data['expenses']:.1f} months"],
    ]
    
    # Write metrics
    for i, (label, value) in enumerate(metrics, start=6):
        ws1[f'A{i}'] = label
        ws1[f'B{i}'] = value
    
    # Adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes

def get_download_link(excel_bytes, filename="financial_report.xlsx"):
    """Generate download link for Excel file"""
    b64 = base64.b64encode(excel_bytes.read()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}" style="background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%); color: white; padding: 12px 24px; border-radius: 10px; text-decoration: none; font-weight: bold; display: inline-block; margin: 10px 0;">📥 Download Financial Report</a>'
    return href

# ========== GRAPH FUNCTIONS ==========
def plot_cash_flow_forecast(revenue, expenses, cash_balance, months=12):
    """Plot cash flow forecast"""
    months_list = list(range(1, months + 1))
    forecast = []
    current_cash = cash_balance
    
    for month in months_list:
        current_cash += (revenue - expenses)
        forecast.append(current_cash)
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=months_list,
        y=forecast,
        mode='lines+markers',
        name='Cash Forecast',
        line=dict(color='#60a5fa', width=3),
        marker=dict(size=8),
        fill='tozeroy',
        fillcolor='rgba(96, 165, 250, 0.1)'
    ))
    
    fig.update_layout(
        title="💰 12-Month Cash Flow Forecast",
        xaxis_title="Months",
        yaxis_title="Cash Balance ($)",
        template="plotly_dark",
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white'),
        hovermode='x unified',
        height=400
    )
    
    return fig

def plot_expense_breakdown(expenses_dict):
    """Plot expense breakdown"""
    fig = go.Figure()
    
    fig.add_trace(go.Pie(
        labels=list(expenses_dict.keys()),
        values=list(expenses_dict.values()),
        hole=0.4,
        marker=dict(colors=['#ef4444', '#f59e0b', '#10b981', '#3b82f6', '#8b5cf6']),
        textinfo='label+percent',
        textposition='outside'
    ))
    
    fig.update_layout(
        title="📊 Expense Breakdown",
        template="plotly_dark",
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white'),
        height=400,
        showlegend=False
    )
    
    return fig

def plot_profit_trend(revenue, expenses):
    """Plot profit trend over time"""
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    profit_trend = []
    
    # Simulate some variation
    for i in range(12):
        month_revenue = revenue * (1 + np.random.uniform(-0.1, 0.2))
        month_expenses = expenses * (1 + np.random.uniform(-0.05, 0.15))
        profit_trend.append(month_revenue - month_expenses)
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=months,
        y=profit_trend,
        mode='lines+markers',
        name='Monthly Profit',
        line=dict(color='#10b981', width=3),
        marker=dict(size=8)
    ))
    
    fig.update_layout(
        title="📈 Monthly Profit Trend",
        xaxis_title="Month",
        yaxis_title="Profit ($)",
        template="plotly_dark",
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white'),
        hovermode='x unified',
        height=400
    )
    
    return fig

def plot_runway_analysis(cash_balance, monthly_expenses):
    """Plot runway analysis"""
    if monthly_expenses == 0:
        return None
    
    scenarios = {
        'Current': cash_balance / monthly_expenses,
        'Reduce Expenses 20%': cash_balance / (monthly_expenses * 0.8),
        'Increase Revenue 20%': (cash_balance + (monthly_expenses * 0.2 * 6)) / monthly_expenses,
        'Both Strategies': cash_balance / (monthly_expenses * 0.8) + 2
    }
    
    fig = go.Figure()
    
    colors = ['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6']
    
    for i, (scenario, runway) in enumerate(scenarios.items()):
        fig.add_trace(go.Bar(
            x=[scenario],
            y=[runway],
            name=scenario,
            marker_color=colors[i],
            text=[f'{runway:.1f} months'],
            textposition='outside'
        ))
    
    fig.add_hline(y=6, line_dash="dash", line_color="green", 
                  annotation_text="6-Month Safety Net", 
                  annotation_position="top right")
    
    fig.update_layout(
        title="🛡️ Runway Analysis - Different Strategies",
        yaxis_title="Months of Runway",
        showlegend=False,
        template="plotly_dark",
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white'),
        height=400
    )
    
    return fig

# ========== SESSION STATE ==========
if "messages" not in st.session_state:
    st.session_state.messages = [
        {"role": "assistant", "content": "👋 **Welcome to UXXCA AI CFO!** I'm your financial co-pilot. I can analyze your finances, generate professional spreadsheets, create interactive graphs, and provide actionable advice."}
    ]

if "financial_data" not in st.session_state:
    st.session_state.financial_data = {
        "revenue": 15000,
        "expenses": 12000,
        "cash_balance": 50000,
        "company_name": "Your Business"
    }

# ========== HELPER FUNCTIONS ==========
def ask_cfo_assistant(question, financial_context=None):
    """Enhanced CFO AI Assistant"""
    if financial_context:
        return f"""
**Analysis of your financial situation:**

**Current Metrics:**
- Monthly Revenue: ${financial_context['revenue']:,.2f}
- Monthly Expenses: ${financial_context['expenses']:,.2f}
- Monthly Profit: ${financial_context['revenue'] - financial_context['expenses']:,.2f}
- Profit Margin: {(financial_context['revenue'] - financial_context['expenses']) / financial_context['revenue'] * 100:.1f}%
- Cash Runway: {financial_context['cash_balance'] / financial_context['expenses']:.1f} months

**Recommendations:**
1. Focus on increasing your profit margin by optimizing expenses
2. Maintain at least 6 months of cash runway for safety
3. Consider reinvesting profits into growth opportunities

**Next Steps:**
Generate detailed reports using the spreadsheet feature below, or explore interactive graphs to visualize your financial health.
"""
    return "I'm here to help with your financial analysis. Please provide your financial data in the sidebar."

# ========== SIDEBAR ==========
with st.sidebar:
    st.markdown("""
    <div style="text-align: center; padding: 1.5rem 0;">
        <h1 style="color: #60a5fa; font-size: 2.2rem; margin: 0;">UXXCA</h1>
        <p style="color: #cbd5e1; font-size: 0.9rem;">AI CFO Assistant</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### 📊 Financial Input")
    
    company_name = st.text_input("**Company Name**", 
                                 value=st.session_state.financial_data.get('company_name', ''),
                                 key="company_name")
    
    revenue = st.number_input("**Monthly Revenue ($)**", 
                             min_value=0, 
                             value=st.session_state.financial_data['revenue'],
                             key="revenue_input")
    
    expenses = st.number_input("**Monthly Expenses ($)**", 
                              min_value=0, 
                              value=st.session_state.financial_data['expenses'],
                              key="expenses_input")
    
    cash_balance = st.number_input("**Cash Balance ($)**", 
                                  min_value=0, 
                                  value=st.session_state.financial_data['cash_balance'],
                                  key="cash_input")
    
    # Expense breakdown inputs
    st.markdown("### 📈 Expense Categories")
    marketing = st.number_input("Marketing ($)", min_value=0, value=int(expenses * 0.2), key="marketing")
    salaries = st.number_input("Salaries ($)", min_value=0, value=int(expenses * 0.4), key="salaries")
    operations = st.number_input("Operations ($)", min_value=0, value=int(expenses * 0.2), key="operations")
    software = st.number_input("Software ($)", min_value=0, value=int(expenses * 0.1), key="software")
    other = st.number_input("Other ($)", min_value=0, value=int(expenses * 0.1), key="other")
    
    if st.button("🔄 Update All Data", type="primary", use_container_width=True):
        st.session_state.financial_data = {
            "revenue": revenue,
            "expenses": expenses,
            "cash_balance": cash_balance,
            "company_name": company_name,
            "expense_breakdown": {
                "Marketing": marketing,
                "Salaries": salaries,
                "Operations": operations,
                "Software": software,
                "Other": other
            }
        }
        st.success("✅ All data updated!")
        st.rerun()

# ========== MAIN INTERFACE ==========
# Header
st.markdown("""
<div class="uxxca-header">
    <h1 style="margin: 0; font-size: 2.8rem;">Financial Clarity at Your Fingertips</h1>
    <p style="margin: 0.8rem 0 0 0; font-size: 1.2rem;">
        AI CFO Assistant • Professional Spreadsheets • Interactive Graphs
    </p>
</div>
""", unsafe_allow_html=True)

# ========== FINANCIAL METRICS ==========
st.markdown("### 📊 Live Financial Dashboard")

col1, col2, col3, col4 = st.columns(4)

profit = revenue - expenses
runway = cash_balance / expenses if expenses > 0 else 0
profit_margin = (profit / revenue * 100) if revenue > 0 else 0

with col1:
    st.metric("Monthly Revenue", f"${revenue:,.0f}")
with col2:
    st.metric("Monthly Expenses", f"${expenses:,.0f}")
with col3:
    st.metric("Monthly Profit", f"${profit:,.0f}", 
              f"{profit_margin:.1f}% margin")
with col4:
    runway_status = "✅" if runway >= 6 else "⚠️" if runway >= 3 else "🚨"
    st.metric("Cash Runway", f"{runway:.1f} months", runway_status)

# ========== SPREADSHEET GENERATOR ==========
st.markdown("---")
st.markdown("### 📈 Generate Professional Spreadsheet")

if st.button("🚀 Generate Excel Financial Report", type="primary", use_container_width=True):
    with st.spinner("Creating professional Excel report..."):
        # Generate Excel file
        excel_file = create_financial_spreadsheet(st.session_state.financial_data)
        
        # Create download link
        st.markdown(get_download_link(excel_file, f"{company_name}_Financial_Report.xlsx"), unsafe_allow_html=True)
        
        st.success("✅ Report generated! Click the button above to download.")

# ========== INTERACTIVE GRAPHS ==========
st.markdown("---")
st.markdown("### 📊 Interactive Financial Graphs")

# Create tabs for different graphs
tab1, tab2, tab3, tab4 = st.tabs(["Cash Flow Forecast", "Expense Breakdown", "Profit Trend", "Runway Analysis"])

with tab1:
    st.markdown('<div class="graph-box">', unsafe_allow_html=True)
    fig1 = plot_cash_flow_forecast(revenue, expenses, cash_balance)
    st.plotly_chart(fig1, use_container_width=True)
    
    st.markdown("**💡 Insights:**")
    st.markdown(f"""
    - Your cash will reach **${cash_balance + (revenue - expenses) * 12:,.0f}** in 12 months
    - Monthly cash flow: **${revenue - expenses:,.0f}**
    - Break-even point: {abs(cash_balance / (revenue - expenses)):.1f} months (if negative cash flow)
    """)
    st.markdown('</div>', unsafe_allow_html=True)

with tab2:
    st.markdown('<div class="graph-box">', unsafe_allow_html=True)
    expense_data = {
        "Marketing": marketing,
        "Salaries": salaries,
        "Operations": operations,
        "Software": software,
        "Other": other
    }
    fig2 = plot_expense_breakdown(expense_data)
    st.plotly_chart(fig2, use_container_width=True)
    
    st.markdown("**💡 Insights:**")
    largest_expense = max(expense_data, key=expense_data.get)
    st.markdown(f"""
    - Largest expense category: **{largest_expense}** (${expense_data[largest_expense]:,.0f})
    - Total expenses: **${expenses:,.0f}**
    - Potential optimization: Focus on reducing the largest expense categories first
    """)
    st.markdown('</div>', unsafe_allow_html=True)

with tab3:
    st.markdown('<div class="graph-box">', unsafe_allow_html=True)
    fig3 = plot_profit_trend(revenue, expenses)
    st.plotly_chart(fig3, use_container_width=True)
    
    st.markdown("**💡 Insights:**")
    st.markdown(f"""
    - Average monthly profit: **${profit:,.0f}**
    - Profit margin target: **20%+** (currently {profit_margin:.1f}%)
    - Seasonality: Consider adjusting for business cycles
    """)
    st.markdown('</div>', unsafe_allow_html=True)

with tab4:
    st.markdown('<div class="graph-box">', unsafe_allow_html=True)
    fig4 = plot_runway_analysis(cash_balance, expenses)
    if fig4:
        st.plotly_chart(fig4, use_container_width=True)
        
        st.markdown("**💡 Recommendations:**")
        if runway < 3:
            st.error("**🚨 CRITICAL:** Runway below 3 months! Immediate action required.")
            st.markdown("1. Cut non-essential expenses immediately")
            st.markdown("2. Accelerate accounts receivable")
            st.markdown("3. Explore emergency funding")
        elif runway < 6:
            st.warning("**⚠️ WARNING:** Runway below 6-month safety net")
            st.markdown("1. Reduce discretionary spending")
            st.markdown("2. Renegotiate vendor contracts")
            st.markdown("3. Delay non-critical hires")
        else:
            st.success("**✅ HEALTHY:** Runway above 6 months")
            st.markdown("1. Consider growth investments")
            st.markdown("2. Build cash reserves")
            st.markdown("3. Explore new opportunities")
    st.markdown('</div>', unsafe_allow_html=True)

# ========== CHAT INTERFACE ==========
st.markdown("---")
st.markdown("### 💬 AI CFO Assistant")

# Display chat
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

# Chat input
if prompt := st.chat_input("💭 Ask your AI CFO about financial strategies, analysis, or report generation..."):
    st.session_state.messages.append({"role": "user", "content": prompt})
    
    with st.chat_message("user"):
        st.markdown(prompt)
    
    with st.chat_message("assistant"):
        with st.spinner("🔍 Analyzing your finances..."):
            response = ask_cfo_assistant(prompt, st.session_state.financial_data)
            st.markdown(response)
            st.session_state.messages.append({"role": "assistant", "content": response})

# ========== QUICK ACTIONS ==========
st.markdown("---")
st.markdown("### 🚀 Quick Actions")

action_cols = st.columns(4)
with action_cols[0]:
    if st.button("📊 Full Analysis", use_container_width=True):
        st.session_state.messages.append({
            "role": "user", 
            "content": "Provide a comprehensive financial analysis with recommendations"
        })
        st.rerun()

with action_cols[1]:
    if st.button("💰 Optimize Expenses", use_container_width=True):
        st.session_state.messages.append({
            "role": "user", 
            "content": f"Suggest ways to optimize my ${expenses:,.0f} in monthly expenses"
        })
        st.rerun()

with action_cols[2]:
    if st.button("📈 Growth Plan", use_container_width=True):
        st.session_state.messages.append({
            "role": "user", 
            "content": f"Create a 6-month growth plan for ${revenue:,.0f} revenue business"
        })
        st.rerun()

with action_cols[3]:
    if st.button("🔄 Update Graphs", use_container_width=True):
        st.rerun()

# ========== FOOTER ==========
st.markdown("---")
st.markdown("""
<div style="text-align: center; padding: 2rem 0;">
    <p style="margin: 0.5rem 0;">
        <strong>UXXCA AI CFO Assistant</strong> • Professional Reports • Interactive Graphs • Actionable Insights
    </p>
    <p style="margin: 0.5rem 0; font-size: 0.9rem; color: #64748b;">
        Generate investor-ready reports and visualize your financial health in real-time
    </p>
</div>
""", unsafe_allow_html=True)
